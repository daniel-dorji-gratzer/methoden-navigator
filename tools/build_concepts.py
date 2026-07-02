#!/usr/bin/env python3
"""Build: erzeugt data/concepts.js aus dem Blatt 'Glossar & Konzepte' der Katalog-XLSX.

Erster Baustein der Tabelle->Daten-Pipeline. Liest das Glossar-/Konzept-Blatt und
schreibt window.CONCEPTS (Schluessel = Begriff-ID).

Aufruf:  python3 tools/build_concepts.py [pfad/zur/Katalog.xlsx]
Standardpfad: ~/Downloads/Methoden_Katalog_v3.xlsx

Hinweis: Texte werden derzeit unveraendert uebernommen. Eine spaetere Stufe kann
hier eine leichte Markdown->HTML-Konvertierung (**fett** -> <b>) einhaengen.
"""
import json, sys, re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
# Kanonische Quelle: neuester Katalog-Export im Projektordner (v4, v5, ...).
SOURCE_DIR = Path.home() / "Documents" / "3_Arbeit_und_Projekte" / "Ψ_Methoden_App"


def newest_catalog():
    cands = sorted(SOURCE_DIR.glob("Methoden_Katalog*.xlsx"),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if not cands:
        sys.exit(f"Kein Methoden_Katalog*.xlsx in {SOURCE_DIR}")
    return cands[0]
HEADER = ("/* Datengetrieben – AUTOMATISCH erzeugt aus dem Katalog-Blatt "
          "'Glossar & Konzepte'.\n   Nicht von Hand editieren. Struktur: data/schema.md */\n")


def split_list(val):
    if not val:
        return []
    parts = re.split(r"[;,]", str(val))
    return [p.strip() for p in parts if p.strip() and p.strip() != "—"]


def main():
    import openpyxl
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else newest_catalog()
    if not xlsx.exists():
        sys.exit(f"Quelldatei nicht gefunden: {xlsx}")
    print(f"Quelle: {xlsx}")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    # Glossar-Blatt robust finden (z. B. 'Glossar & Konzepte' oder 'Glossar_Konzepte')
    sheet = next((s for s in wb.sheetnames
                  if "glossar" in s.lower() or "konzept" in s.lower()), None)
    if not sheet:
        sys.exit(f"Kein Glossar-/Konzept-Blatt in {xlsx.name} gefunden")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h else "" for h in rows[0]]

    def col(name):
        nl = name.lower()
        for i, h in enumerate(hdr):       # exakte Übereinstimmung zuerst
            if h.lower() == nl:
                return i
        for i, h in enumerate(hdr):       # sonst Teilstring
            if nl in h.lower():
                return i
        raise KeyError(f"Spalte '{name}' nicht gefunden in {hdr}")
    c = dict(id=col("Begriff-ID"), term=col("Begriff"), aliases=col("Synonyme"),
             cat=col("Kategorie"), short=col("Kurzdefinition"), long=col("Erklärung"),
             rel=col("Verknüpfte"), status=col("Status"))

    concepts = {}
    for r in rows[1:]:
        cid = r[c["id"]]
        if not cid:
            continue
        cid = str(cid).strip()
        entry = {
            "term": (r[c["term"]] or "").strip(),
            "aliases": split_list(r[c["aliases"]]),
            "category": (r[c["cat"]] or "").strip(),
            "short": (r[c["short"]] or "").strip(),
            "long": (r[c["long"]] or "").strip(),
            "related": split_list(r[c["rel"]]),
            "status": (r[c["status"]] or "").strip(),
        }
        concepts[cid] = entry

    out = ROOT / "data" / "concepts.js"
    out.write_text(HEADER + "window.CONCEPTS = " +
                   json.dumps(concepts, ensure_ascii=False, indent=2) + ";\n",
                   encoding="utf-8")
    print(f"OK: {len(concepts)} Konzepte -> {out.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
